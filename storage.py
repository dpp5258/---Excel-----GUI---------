"""Excel 存储层 — 文件读写、初始化、备份、校验"""

import os
import sys
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook


def get_app_dir() -> str:
    """返回 .exe 或 .py 所在的目录路径。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


class ExcelStorage:
    FILE_NAME = os.path.join(get_app_dir(), "任务清单.xlsx")
    BACKUP_DIR = os.path.join(get_app_dir(), "backup")
    MAX_BACKUPS = 10
    SHEET_TODO = "待办任务"
    SHEET_DONE = "已完成任务"
    SHEET_ARCHIVE = "归档任务"
    HEADERS = [
        "ID", "任务标题", "任务内容", "创建时间", "截止时间",
        "优先级", "备注", "完成状态", "完成情况",
    ]

    # ── 公开接口 ──────────────────────────────────────

    def init_file(self) -> None:
        """初始化文件：不存在则创建，表头损坏则重建。"""
        os.makedirs(self.BACKUP_DIR, exist_ok=True)
        if not os.path.exists(self.FILE_NAME):
            self._create_fresh()
            return
        if not self.validate_headers():
            self._rebuild_file()

    def read_todo(self) -> list[list]:
        return self._read_sheet(self.SHEET_TODO)

    def read_done(self) -> list[list]:
        return self._read_sheet(self.SHEET_DONE)

    def read_archive(self) -> list[list]:
        return self._read_sheet(self.SHEET_ARCHIVE)

    def read_all(self) -> tuple[list[list], list[list], list[list]]:
        return (self.read_todo(), self.read_done(), self.read_archive())

    def write_all(
        self,
        todo_data: list[list],
        done_data: list[list],
        archive_data: list[list],
    ) -> None:
        """全量写入三张Sheet（写入前自动备份）。"""
        self.backup()
        wb = self._open_workbook()
        try:
            self._write_sheet(wb, self.SHEET_TODO, todo_data)
            self._write_sheet(wb, self.SHEET_DONE, done_data)
            self._write_sheet(wb, self.SHEET_ARCHIVE, archive_data)
            wb.save(self.FILE_NAME)
        except PermissionError:
            raise PermissionError(
                "文件被占用，请关闭 Excel/WPS 后重试"
            )
        finally:
            wb.close()

    def get_max_id(self) -> int:
        """遍历三张Sheet，返回最大ID。无数据返回0。"""
        max_id = 0
        for sheet_name in (self.SHEET_TODO, self.SHEET_DONE, self.SHEET_ARCHIVE):
            rows = self._read_sheet(sheet_name)
            for row in rows:
                try:
                    val = int(row[0])
                    if val > max_id:
                        max_id = val
                except (ValueError, IndexError):
                    continue
        return max_id

    def backup(self) -> None:
        """复制 Excel 到 backup/ 目录（带时间戳），保留最近 MAX_BACKUPS 份。"""
        if not os.path.exists(self.FILE_NAME):
            return
        os.makedirs(self.BACKUP_DIR, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name, ext = os.path.splitext(os.path.basename(self.FILE_NAME))
        backup_name = f"{self.BACKUP_DIR}/{name}_{timestamp}{ext}"
        shutil.copy2(self.FILE_NAME, backup_name)
        self._rotate_backups()

    def validate_headers(self) -> bool:
        """校验三张Sheet表头是否与 HEADERS 一致。"""
        try:
            wb = load_workbook(self.FILE_NAME, read_only=True)
        except Exception:
            return False
        try:
            for sheet_name in (self.SHEET_TODO, self.SHEET_DONE, self.SHEET_ARCHIVE):
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return False
                ws = wb[sheet_name]
                row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if row != self.HEADERS:
                    wb.close()
                    return False
        finally:
            wb.close()
        return True

    # ── 私有方法 ──────────────────────────────────────

    def _open_workbook(self) -> Workbook:
        """打开工作簿（文件不存在则创建）。"""
        if os.path.exists(self.FILE_NAME):
            return load_workbook(self.FILE_NAME)
        return self._create_fresh()

    def _create_fresh(self) -> Workbook:
        """创建全新工作簿，包含三Sheet表头。"""
        wb = Workbook()
        wb.active.title = self.SHEET_TODO
        ws_todo = wb.active
        ws_todo.append(self.HEADERS)
        wb.create_sheet(self.SHEET_DONE).append(self.HEADERS)
        wb.create_sheet(self.SHEET_ARCHIVE).append(self.HEADERS)
        wb.save(self.FILE_NAME)
        return wb

    def _rebuild_file(self) -> None:
        """表头损坏时：备份旧数据 → 重建文件 → 恢复有效数据行。"""
        old_rows = {s: [] for s in (self.SHEET_TODO, self.SHEET_DONE, self.SHEET_ARCHIVE)}
        try:
            old_wb = load_workbook(self.FILE_NAME)
            for sheet_name in old_rows:
                if sheet_name in old_wb.sheetnames:
                    ws = old_wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            old_rows[sheet_name].append(list(row))
            old_wb.close()
        except Exception:
            pass

        wb = self._create_fresh()
        wb.close()
        self.write_all(old_rows[self.SHEET_TODO],
                       old_rows[self.SHEET_DONE],
                       old_rows[self.SHEET_ARCHIVE])

    def _read_sheet(self, sheet_name: str) -> list[list]:
        """读取指定Sheet的所有数据行（不含表头）。"""
        if not os.path.exists(self.FILE_NAME):
            return []
        wb = load_workbook(self.FILE_NAME, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append(list(row))
            return rows
        finally:
            wb.close()

    def _write_sheet(self, wb: Workbook, sheet_name: str, data: list[list]) -> None:
        """清空指定Sheet，写入表头和数据行。"""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(sheet_name)
        ws.append(self.HEADERS)
        for row in data:
            ws.append(row)

    def _rotate_backups(self) -> None:
        """删除旧备份，只保留最近 MAX_BACKUPS 份。"""
        prefix = os.path.splitext(os.path.basename(self.FILE_NAME))[0]
        files = sorted([
            f for f in os.listdir(self.BACKUP_DIR)
            if f.startswith(prefix) and f.endswith(".xlsx")
        ])
        while len(files) > self.MAX_BACKUPS:
            os.remove(os.path.join(self.BACKUP_DIR, files.pop(0)))
